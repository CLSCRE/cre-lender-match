#!/usr/bin/env python3
"""
CRE Lender Match Tool
=====================
Finds banks and credit unions actively lending on commercial real estate
in a specific market, ranked by actual origination activity.

Data Sources:
  - HMDA (CFPB) -- Actual multifamily loan origination records
  - FDIC BankFind -- Bank financial data and CRE portfolio concentrations
  - NCUA -- Credit union financial data and CRE portfolio concentrations
  - ATTOM Data API -- County recorder data for all CRE types (requires API key)

Usage:
  python cre_lender_match.py --city Miami --state FL --min-loan 2000000 --max-loan 10000000
  python cre_lender_match.py --city Phoenix --state AZ --years 2022,2023,2024
  python cre_lender_match.py --county 12086 --min-loan 1000000
  python cre_lender_match.py --state TX --min-loan 5000000 --max-loan 25000000
"""

import argparse
import csv
import io
import json
import os
import sys
import time
import zipfile
from collections import defaultdict
from datetime import datetime
from urllib.parse import urlencode, quote

try:
    import requests
except ImportError:
    print("ERROR: 'requests' package required. Run: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: 'openpyxl' package required. Run: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None  # Optional -- will use simple print progress


# ════════════════════════════════════════════════════════════════════════════════
# Configuration
# ════════════════════════════════════════════════════════════════════════════════

HMDA_BASE = "https://ffiec.cfpb.gov/v2/data-browser-api/view"
FDIC_BASE = "https://api.fdic.gov/banks"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
FCC_AREA_URL = "https://geo.fcc.gov/api/census/area"

DEFAULT_YEARS = [2022, 2023, 2024]
MULTIFAMILY_UNITS = "5-24,25-49,50-99,100-149,>149"

ATTOM_API_KEY = os.environ.get("ATTOM_API_KEY", "")

# NCUA API endpoints (credit union data)
NCUA_SEARCH_URL = "https://mapping.ncua.gov/api/Search/GetSearchLocations"
NCUA_DETAILS_URL = "https://mapping.ncua.gov/api/CreditUnionDetails/GetCreditUnionDetails"
NCUA_CYCLE_URL = "https://mapping.ncua.gov/api/DataQuery/GetCurrentCycle"
NCUA_DATA_URL = "https://www.ncua.gov/files/publications/analysis/call-report-data-{year}-{month:02d}.zip"
NCUA_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ncua_data")

# Request settings
REQUEST_TIMEOUT = 120
FDIC_SEARCH_DELAY = 0.3  # seconds between FDIC API calls to be respectful
NCUA_SEARCH_DELAY = 0.3  # seconds between NCUA API calls


# ════════════════════════════════════════════════════════════════════════════════
# Geocoding: City + State -> County FIPS Code
# ════════════════════════════════════════════════════════════════════════════════

def geocode_city_to_fips(city: str, state: str) -> dict:
    """
    Convert city + state to county FIPS code.
    Uses Nominatim (city -> lat/lon) then FCC Area API (lat/lon -> county FIPS).
    Returns dict with keys: county_fips, county_name
    """
    print(f"  Geocoding '{city}, {state}'...")

    # Step 1: City -> lat/lon via Nominatim (OpenStreetMap)
    params = {
        "city": city,
        "state": state,
        "country": "US",
        "format": "json",
        "limit": "1",
    }
    url = f"{NOMINATIM_URL}?{urlencode(params)}"
    try:
        resp = requests.get(url, timeout=15, headers={"User-Agent": "CRE-Lender-Match/1.0"})
        resp.raise_for_status()
        results = resp.json()
    except Exception as e:
        print(f"  WARNING: Nominatim geocoding failed: {e}")
        return None

    if not results:
        print(f"  WARNING: No geocoding match for '{city}, {state}'")
        return None

    lat = results[0]["lat"]
    lon = results[0]["lon"]

    # Step 2: lat/lon -> county FIPS via FCC Area API
    fcc_url = f"{FCC_AREA_URL}?lat={lat}&lon={lon}&format=json"
    try:
        resp2 = requests.get(fcc_url, timeout=15)
        resp2.raise_for_status()
        fcc_data = resp2.json()
    except Exception as e:
        print(f"  WARNING: FCC area lookup failed: {e}")
        return None

    fcc_results = fcc_data.get("results", [])
    if not fcc_results:
        print(f"  WARNING: No county found for coordinates ({lat}, {lon})")
        return None

    county_fips = fcc_results[0].get("county_fips", "")
    county_name = fcc_results[0].get("county_name", "").replace(" County", "")

    if not county_fips:
        print(f"  WARNING: No FIPS code returned")
        return None

    print(f"  Found: {county_name} County (FIPS: {county_fips})")
    return {
        "county_fips": county_fips,
        "county_name": county_name,
    }


# ════════════════════════════════════════════════════════════════════════════════
# HMDA Client: Loan Origination Data
# ════════════════════════════════════════════════════════════════════════════════

def fetch_hmda_filers(years: list) -> dict:
    """
    Fetch the LEI -> institution name mapping from HMDA filers endpoint.
    Returns dict: {lei: name}
    """
    print(f"\n  Fetching HMDA filer directory ({', '.join(str(y) for y in years)})...")
    filers = {}

    for year in years:
        url = f"{HMDA_BASE}/filers?years={year}"
        try:
            resp = requests.get(url, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            data = resp.json()
            for inst in data.get("institutions", []):
                lei = inst.get("lei", "")
                name = inst.get("name", "")
                if lei and name:
                    filers[lei] = name
        except Exception as e:
            print(f"  WARNING: Failed to fetch filers for {year}: {e}")

    print(f"  Loaded {len(filers):,} institution names")
    return filers


def fetch_hmda_originations(geo_type: str, geo_code: str, years: list) -> list:
    """
    Download loan-level multifamily origination records from HMDA.

    Args:
        geo_type: 'counties', 'msamds', or 'states'
        geo_code: FIPS code, MSA code, or state abbreviation
        years: list of years to query

    Returns: list of loan record dicts
    """
    years_str = ",".join(str(y) for y in years)
    params = {
        "years": years_str,
        "actions_taken": "1",  # Originations only
        "total_units": MULTIFAMILY_UNITS,  # 5+ units = multifamily
    }
    params[geo_type] = geo_code

    # Build URL -- need special handling for >149 encoding
    param_str = urlencode(params, safe=",")
    param_str = param_str.replace("%3E", ">")  # HMDA API expects literal >
    url = f"{HMDA_BASE}/csv?{param_str}"

    geo_label = f"{geo_type}={geo_code}"
    print(f"\n  Downloading HMDA originations ({geo_label}, {years_str})...")
    print(f"  This may take a moment for large geographies...")

    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT * 2, stream=True)
        resp.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"  ERROR: HMDA API returned {e.response.status_code}")
        if e.response.status_code == 400:
            print(f"  Check that your geography code is valid: {geo_code}")
        return []
    except Exception as e:
        print(f"  ERROR: Failed to download HMDA data: {e}")
        return []

    # Parse CSV
    content = resp.text
    if not content.strip():
        print("  No data returned from HMDA")
        return []

    reader = csv.DictReader(io.StringIO(content))
    loans = []
    for row in reader:
        loans.append(row)

    print(f"  Downloaded {len(loans):,} multifamily origination records")
    return loans


# ════════════════════════════════════════════════════════════════════════════════
# FDIC Client: Bank Financial Data
# ════════════════════════════════════════════════════════════════════════════════

def search_fdic_institution(name: str) -> dict:
    """
    Search FDIC BankFind for a bank by name using wildcard filters.
    Returns institution dict or None.
    """
    # Clean the name -- remove punctuation, common suffixes
    clean = name.replace(",", " ").replace(".", " ").replace("-", " ").strip()
    skip_words = {"THE", "OF", "AND", "INC", "LLC", "NA", "NATIONAL", "ASSOCIATION",
                  "BANK", "FSB", "SSB", "COMPANY", "CORP", "CORPORATION", "GROUP"}

    # Get distinctive keywords for wildcard search
    words = [w for w in clean.split() if w.upper() not in skip_words and len(w) > 1]

    if not words:
        # Fallback: just use the first word of the original name
        words = clean.split()[:1]

    # Try progressively broader searches
    fields = "CERT,NAME,CITY,STALP,ASSET,DEP,EQ,WEBADDR,OFFDOM,SPECGRPN,ADDRESS,ZIP,ACTIVE"

    for attempt in range(min(3, len(words)), 0, -1):
        # Build filter: NAME:word1* AND NAME:*word2*
        keyword = words[0]
        name_filter = f"ACTIVE:1 AND NAME:{keyword}*"
        params = {
            "filters": name_filter,
            "fields": fields,
            "limit": "10",
            "sort_by": "ASSET",
            "sort_order": "DESC",
        }
        url = f"{FDIC_BASE}/institutions?{urlencode(params)}"

        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            results = data.get("data", [])

            if results:
                # Score each result against the original name for best match
                best_match = None
                best_score = 0
                name_upper = name.upper()

                for r in results:
                    rdata = r.get("data", {})
                    rname = (rdata.get("NAME", "") or "").upper()
                    # Simple scoring: count matching words
                    score = sum(1 for w in words if w.upper() in rname)
                    # Bonus for exact-ish match
                    if name_upper.split(",")[0].strip() in rname or rname in name_upper:
                        score += 10
                    if score > best_score:
                        best_score = score
                        best_match = rdata

                if best_match and best_score >= 1:
                    return best_match
        except Exception:
            pass

        # Try next keyword if first didn't work
        if len(words) > 1:
            words = words[1:]

    return None


def fetch_fdic_financials(cert: int) -> dict:
    """
    Fetch CRE financial data for a bank from FDIC Call Reports.
    Returns dict with CRE loan fields.
    """
    fields = (
        "REPDTE,CERT,ASSET,EQ,LNRE,LNRECONS,LNRECNFM,LNRECNOT,"
        "LNREMULT,LNRENRES,LNRENROW,LNRENROT,LNREAG,LNRERES,"
        "DEP,NETINC,ROA,ROE,LNLSNET,LNCI,NUMEMP"
    )
    params = {
        "filters": f"CERT:{cert}",
        "fields": fields,
        "sort_by": "REPDTE",
        "sort_order": "DESC",
        "limit": "1",
    }
    url = f"{FDIC_BASE}/financials?{urlencode(params)}"

    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        results = data.get("data", [])
        if results:
            return results[0].get("data", {})
    except Exception:
        pass

    return None


def compute_cre_concentration(financials: dict) -> dict:
    """
    Compute CRE concentration metrics from FDIC financial data.
    All FDIC values are in thousands of dollars.
    """
    if not financials:
        return {}

    eq = financials.get("EQ", 0) or 0
    asset = financials.get("ASSET", 0) or 0
    construction = financials.get("LNRECONS", 0) or 0
    multifamily = financials.get("LNREMULT", 0) or 0
    nonfarm_nonres = financials.get("LNRENRES", 0) or 0
    nonfarm_nonocc = financials.get("LNRENROT", 0) or 0
    nonfarm_ownocc = financials.get("LNRENROW", 0) or 0
    total_re = financials.get("LNRE", 0) or 0

    # Regulatory CRE = construction + multifamily + non-owner-occupied nonfarm nonres
    cre_total = construction + multifamily + nonfarm_nonocc

    # CRE as % of total capital (regulatory threshold is 300%)
    cre_to_capital_pct = (cre_total / eq * 100) if eq > 0 else 0

    # CRE as % of total assets
    cre_to_assets_pct = (cre_total / asset * 100) if asset > 0 else 0

    # Multifamily as % of total RE loans
    mf_pct = (multifamily / total_re * 100) if total_re > 0 else 0

    return {
        "cre_total_k": cre_total,
        "cre_to_capital_pct": round(cre_to_capital_pct, 1),
        "cre_to_assets_pct": round(cre_to_assets_pct, 1),
        "multifamily_k": multifamily,
        "multifamily_pct_of_re": round(mf_pct, 1),
        "construction_k": construction,
        "nonfarm_nonres_k": nonfarm_nonres,
        "nonfarm_nonocc_k": nonfarm_nonocc,
        "nonfarm_ownocc_k": nonfarm_ownocc,
        "total_assets_k": asset,
        "total_equity_k": eq,
        "total_re_loans_k": total_re,
        "report_date": financials.get("REPDTE", ""),
    }


# CRE property type to FDIC/NCUA field mapping
# Call reports only expose 5 buckets (construction, multifamily, non-owner-occ, owner-occ, 1-4 family).
# Niche product types (student housing, hotel, self-storage, etc.) roll up into the closest bucket;
# the specialty-lender overlay in specialty_lenders.json adds niche agency/CMBS/SBA/HUD shops on top.
CRE_TYPE_MAP = {
    "all_cre":              {"label": "All CRE",                      "fdic_fields": ["LNRECONS", "LNREMULT", "LNRENROT"], "ncua_fields": ["construction", "multifamily", "nonocc_nonfarm"]},
    "multifamily":          {"label": "Multifamily (5+ units)",       "fdic_fields": ["LNREMULT"],                         "ncua_fields": ["multifamily"]},
    "student_housing":      {"label": "Student Housing",              "fdic_fields": ["LNREMULT"],                         "ncua_fields": ["multifamily"]},
    "affordable_housing":   {"label": "Affordable Housing (LIHTC)",   "fdic_fields": ["LNREMULT"],                         "ncua_fields": ["multifamily"]},
    "senior_housing":       {"label": "Senior / Independent Living",  "fdic_fields": ["LNREMULT"],                         "ncua_fields": ["multifamily"]},
    "manufactured_housing": {"label": "Manufactured Housing Park",    "fdic_fields": ["LNREMULT"],                         "ncua_fields": ["multifamily"]},
    "hospitality":          {"label": "Hospitality / Hotel",          "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "office":               {"label": "Office",                       "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "retail":               {"label": "Retail / Shopping Center",     "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "industrial":           {"label": "Industrial / Warehouse",       "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "self_storage":         {"label": "Self-Storage",                 "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "healthcare":           {"label": "Healthcare / Medical Office",  "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "senior_care":          {"label": "Senior Care / SNF / ALF",      "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "mixed_use":            {"label": "Mixed-Use",                    "fdic_fields": ["LNRENROT", "LNREMULT"],             "ncua_fields": ["nonocc_nonfarm", "multifamily"]},
    "special_purpose":      {"label": "Special Purpose",              "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "non_res":              {"label": "Non-Residential Investment",   "fdic_fields": ["LNRENROT"],                         "ncua_fields": ["nonocc_nonfarm"]},
    "construction":         {"label": "Construction & Development",   "fdic_fields": ["LNRECONS"],                         "ncua_fields": ["construction", "construction_commercial"]},
    "land":                 {"label": "Land / Development Site",      "fdic_fields": ["LNRECONS"],                         "ncua_fields": ["construction", "construction_commercial"]},
    "owner_occ":             {"label": "Owner-Occupied CRE",          "fdic_fields": ["LNRENROW"],                         "ncua_fields": ["owner_occ_nonfarm"]},
    "sba_owner_user":       {"label": "SBA Owner-User (504 / 7a)",    "fdic_fields": ["LNRENROW"],                         "ncua_fields": ["owner_occ_nonfarm"]},
}


def fetch_state_cre_lenders(state: str, property_type: str = "all_cre",
                            min_assets_m: int = 50, max_results: int = 200,
                            ncua_data: dict = None) -> list:
    """
    Find banks and credit unions with CRE portfolios in a state.
    Uses FDIC call report data (banks) and NCUA call report data (credit unions).
    Returns ranked list of lenders sorted by portfolio size in the selected CRE type.
    """
    cre_config = CRE_TYPE_MAP.get(property_type, CRE_TYPE_MAP["all_cre"])
    fdic_sum_fields = cre_config["fdic_fields"]
    ncua_sum_fields = cre_config["ncua_fields"]

    print(f"\n  Searching for {cre_config['label']} lenders in {state}...")

    # ── Step 1: Get FDIC banks in the state ──
    min_assets_k = min_assets_m * 1000  # FDIC uses thousands
    fields = (
        "CERT,NAME,CITY,STALP,ASSET,DEP,EQ,WEBADDR,ADDRESS,ZIP,OFFDOM,SPECGRPN"
    )
    inst_params = {
        "filters": f"STALP:{state} AND ACTIVE:1 AND ASSET:[{min_assets_k} TO *]",
        "fields": fields,
        "limit": str(max_results),
        "sort_by": "ASSET",
        "sort_order": "DESC",
    }

    print(f"  Fetching FDIC institutions in {state} (assets > ${min_assets_m}M)...")
    try:
        resp = requests.get(f"{FDIC_BASE}/institutions", params=inst_params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        institutions = [r.get("data", {}) for r in data.get("data", [])]
    except Exception as e:
        print(f"  ERROR: Failed to fetch FDIC institutions: {e}")
        institutions = []

    print(f"  Found {len(institutions)} banks")

    # ── Step 2: Fetch CRE financials for each bank ──
    fin_fields = (
        "REPDTE,CERT,ASSET,EQ,LNRE,LNRECONS,LNRECNFM,LNRECNOT,"
        "LNREMULT,LNRENRES,LNRENROW,LNRENROT,DEP,NETINC,ROA,ROE,LNCI"
    )

    rankings = []
    count = len(institutions)
    print(f"  Fetching CRE portfolio data for {count} banks...")
    progress = tqdm(range(count), desc="  FDIC financials") if tqdm else range(count)

    for i in progress:
        inst = institutions[i]
        cert = inst.get("CERT")
        if not cert:
            continue

        # Fetch financials
        try:
            fin_resp = requests.get(f"{FDIC_BASE}/financials", params={
                "filters": f"CERT:{cert}",
                "fields": fin_fields,
                "sort_by": "REPDTE",
                "sort_order": "DESC",
                "limit": "1",
            }, timeout=15)
            fin_resp.raise_for_status()
            fin_data = fin_resp.json()
            fin_results = fin_data.get("data", [])
            fin = fin_results[0].get("data", {}) if fin_results else {}
        except Exception:
            fin = {}

        if not fin:
            time.sleep(0.1)
            continue

        # Compute CRE portfolio for the selected type
        type_portfolio = sum((fin.get(f, 0) or 0) for f in fdic_sum_fields)

        if type_portfolio <= 0:
            time.sleep(0.1)
            continue

        cre = compute_cre_concentration(fin)
        asset_m = round((inst.get("ASSET", 0) or 0) / 1000, 1)

        rankings.append({
            "name": inst.get("NAME", ""),
            "institution_type": "Bank",
            "fdic_match": True,
            "ncua_match": False,
            "fdic_cert": cert,
            "city": inst.get("CITY", ""),
            "state": inst.get("STALP", ""),
            "address": inst.get("ADDRESS", ""),
            "zip": inst.get("ZIP", ""),
            "website": inst.get("WEBADDR", ""),
            "total_assets_m": asset_m,
            "offices": inst.get("OFFDOM", ""),
            "specialization": inst.get("SPECGRPN", ""),
            "type_portfolio_k": type_portfolio,
            "type_portfolio_m": round(type_portfolio / 1000, 1),
            **cre,
        })

        if not tqdm and (i + 1) % 25 == 0:
            print(f"    ...processed {i + 1}/{count}")

        time.sleep(0.15)

    # ── Step 3: Add credit unions from cached NCUA data ──
    if ncua_data:
        cu_count = 0
        for cu_num, cu_info in ncua_data.items():
            cu_state = (cu_info.get("state") or "").upper()
            if cu_state != state.upper():
                continue

            # Compute portfolio for selected type
            type_portfolio = 0
            for field in ncua_sum_fields:
                type_portfolio += (cu_info.get(field, 0) or 0)

            if type_portfolio <= 0:
                continue

            total_assets = cu_info.get("total_assets", 0) or 0
            if total_assets < min_assets_m * 1_000_000:
                continue

            cu_cre = compute_cu_cre_concentration(cu_num, ncua_data)
            asset_m = round(total_assets / 1_000_000, 1)

            rankings.append({
                "name": cu_info.get("cu_name", f"CU #{cu_num}"),
                "institution_type": "Credit Union",
                "fdic_match": False,
                "ncua_match": True,
                "ncua_charter": int(cu_num),
                "city": cu_info.get("city", ""),
                "state": cu_info.get("state", ""),
                "address": cu_info.get("address", ""),
                "zip": cu_info.get("zip", ""),
                "website": cu_info.get("website", ""),
                "phone": cu_info.get("phone", ""),
                "ceo": cu_info.get("ceo", ""),
                "total_assets_m": asset_m,
                "type_portfolio_k": type_portfolio,
                "type_portfolio_m": round(type_portfolio / 1000, 1) if type_portfolio >= 1000 else round(type_portfolio / 1000, 3),
                **cu_cre,
            })
            cu_count += 1

        print(f"  Added {cu_count} credit unions from NCUA data")

    # ── Step 4: Sort by type portfolio size ──
    rankings.sort(key=lambda x: x.get("type_portfolio_k", 0), reverse=True)

    for i, r in enumerate(rankings, 1):
        r["rank"] = i
        # Fill in deal_count etc. for compatibility with output functions
        r.setdefault("deal_count", 0)
        r.setdefault("total_volume", r.get("type_portfolio_k", 0) * 1000)
        r.setdefault("avg_loan_size", 0)
        r.setdefault("min_loan_originated", 0)
        r.setdefault("max_loan_originated", 0)
        r.setdefault("years_active", [])

    banks = sum(1 for r in rankings if r["institution_type"] == "Bank")
    cus = sum(1 for r in rankings if r["institution_type"] == "Credit Union")
    print(f"  Found {len(rankings)} CRE lenders ({banks} banks, {cus} credit unions)")

    return rankings


# ════════════════════════════════════════════════════════════════════════════════
# NCUA Client: Credit Union Data
# ════════════════════════════════════════════════════════════════════════════════

def get_ncua_current_cycle() -> tuple:
    """Get the most recent NCUA call report cycle date. Returns (year, month)."""
    try:
        resp = requests.get(NCUA_CYCLE_URL, timeout=15)
        resp.raise_for_status()
        # Response is a date string like "2025-09-30T00:00:00"
        date_str = resp.json().strip('"')
        dt = datetime.fromisoformat(date_str.replace("Z", ""))
        return (dt.year, dt.month)
    except Exception as e:
        print(f"  WARNING: Could not get NCUA cycle date: {e}")
        # Fall back to a reasonable default
        now = datetime.now()
        # NCUA reports quarterly: 03, 06, 09, 12
        quarter_month = ((now.month - 1) // 3) * 3
        if quarter_month == 0:
            return (now.year - 1, 12)
        return (now.year, quarter_month)


def download_ncua_call_reports(year: int, month: int) -> dict:
    """
    Download NCUA quarterly call report data and parse CRE-relevant fields.
    Caches parsed data as JSON for fast subsequent lookups.
    Returns dict: {cu_number: {field: value, ...}}
    """
    os.makedirs(NCUA_CACHE_DIR, exist_ok=True)
    cache_file = os.path.join(NCUA_CACHE_DIR, f"cre_data_{year}_{month:02d}.json")

    # Check cache
    if os.path.exists(cache_file):
        print(f"  Loading cached NCUA data ({year}-{month:02d})...")
        with open(cache_file, "r") as f:
            return json.load(f)

    # Download ZIP
    url = NCUA_DATA_URL.format(year=year, month=month)
    print(f"  Downloading NCUA call report data ({year}-{month:02d})...")
    print(f"  URL: {url}")
    print(f"  This is a one-time download (~30MB), cached for future runs...")

    try:
        resp = requests.get(url, timeout=300, stream=True)
        resp.raise_for_status()
    except Exception as e:
        print(f"  WARNING: Failed to download NCUA data: {e}")
        # Try previous quarter
        prev_month = month - 3
        prev_year = year
        if prev_month <= 0:
            prev_month += 12
            prev_year -= 1
        alt_url = NCUA_DATA_URL.format(year=prev_year, month=prev_month)
        print(f"  Trying previous quarter ({prev_year}-{prev_month:02d})...")
        try:
            resp = requests.get(alt_url, timeout=300, stream=True)
            resp.raise_for_status()
        except Exception as e2:
            print(f"  WARNING: Also failed: {e2}")
            return {}

    # Save ZIP to temp file
    zip_path = os.path.join(NCUA_CACHE_DIR, "temp_callreport.zip")
    total_size = int(resp.headers.get("content-length", 0))
    downloaded = 0
    with open(zip_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=65536):
            f.write(chunk)
            downloaded += len(chunk)
            if total_size > 0 and downloaded % (1024 * 1024) < 65536:
                pct = downloaded / total_size * 100
                print(f"    Downloaded {downloaded // (1024*1024)}MB / {total_size // (1024*1024)}MB ({pct:.0f}%)")

    print(f"  Parsing call report data...")

    cu_data = {}

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            file_list = zf.namelist()

            # Parse FOICU.txt - credit union profile data (name, city, state)
            foicu_files = [f for f in file_list if f.upper().endswith("FOICU.TXT")]
            if foicu_files:
                print(f"    Reading {foicu_files[0]} (profiles)...")
                with zf.open(foicu_files[0]) as f:
                    content = f.read()
                    text = content.decode("latin-1")
                    reader = csv.DictReader(io.StringIO(text))
                    if reader.fieldnames:
                        reader.fieldnames = [h.strip() for h in reader.fieldnames]

                    # Print headers once to help discover column names
                    if reader.fieldnames:
                        print(f"    FOICU.txt headers ({len(reader.fieldnames)}): {reader.fieldnames[:30]}...")

                    for row in reader:
                        cu_num = (row.get("CU_NUMBER") or "").strip()
                        if not cu_num:
                            continue
                        # Phone: try common NCUA column names
                        phone = (row.get("PHONE") or row.get("PHONENUMBER")
                                 or row.get("PHONE_NUMBER") or "").strip()
                        # CEO name
                        ceo = (row.get("CEO") or row.get("CEO_NAME")
                               or row.get("CEONAME") or "").strip()
                        # Street address
                        address = (row.get("STREET") or row.get("ADDRESS")
                                   or row.get("PHYSICALADDRESSLINE1")
                                   or row.get("PHYSICAL_ADDRESS_LINE1") or "").strip()
                        # Zip code
                        zipcode = (row.get("ZIP_CODE") or row.get("ZIPCODE")
                                   or row.get("ZIP") or "").strip()
                        # Website
                        website = (row.get("URL") or row.get("WEBSITE")
                                   or row.get("SITE_URL") or row.get("WEBSITE_URL") or "").strip()

                        cu_data[cu_num] = {
                            "cu_name": (row.get("CU_NAME") or "").strip(),
                            "city": (row.get("CITY") or row.get("PHYS_CITY") or "").strip(),
                            "state": (row.get("STATE") or row.get("PHYS_STATE") or "").strip(),
                            "phone": phone,
                            "ceo": ceo,
                            "address": address,
                            "zip": zipcode,
                            "website": website,
                        }

            # Parse FS220.txt - main balance sheet (total assets, total loans, total RE)
            fs220_files = [f for f in file_list if f.upper().endswith("FS220.TXT") and "FS220" in f.upper() and "FS220A" not in f.upper() and "FS220B" not in f.upper() and "FS220C" not in f.upper() and "FS220D" not in f.upper() and "FS220G" not in f.upper() and "FS220H" not in f.upper() and "FS220I" not in f.upper() and "FS220J" not in f.upper() and "FS220K" not in f.upper() and "FS220L" not in f.upper() and "FS220M" not in f.upper() and "FS220N" not in f.upper() and "FS220P" not in f.upper() and "FS220Q" not in f.upper() and "FS220R" not in f.upper() and "FS220S" not in f.upper()]
            if not fs220_files:
                # Try broader match
                fs220_files = [f for f in file_list if "FS220" in f.upper() and f.upper().replace("\\", "/").split("/")[-1] == "FS220.TXT"]
            if not fs220_files:
                fs220_files = [f for f in file_list if f.upper().endswith("FS220.TXT")]

            if fs220_files:
                print(f"    Reading {fs220_files[0]}...")
                with zf.open(fs220_files[0]) as f:
                    content = f.read()
                    # Try different encodings
                    for encoding in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            text = content.decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    else:
                        text = content.decode("latin-1")

                    reader = csv.DictReader(io.StringIO(text))
                    # Normalize header names (strip whitespace)
                    if reader.fieldnames:
                        reader.fieldnames = [h.strip() for h in reader.fieldnames]

                    for row in reader:
                        cu_num = (row.get("CU_NUMBER") or row.get("cu_number") or "").strip()
                        if not cu_num:
                            continue
                        if cu_num not in cu_data:
                            cu_data[cu_num] = {}
                        cu_data[cu_num].update({
                            "total_assets": _safe_int(row.get("ACCT_010", 0)),
                            "total_loans": _safe_int(row.get("ACCT_025B", 0)),
                            "total_1st_mtg_re": _safe_int(row.get("ACCT_703", 0)),
                            "total_other_re": _safe_int(row.get("ACCT_386", 0)),
                            "net_mbl": _safe_int(row.get("ACCT_400A", 0)),
                            "construction": _safe_int(row.get("ACCT_143B", 0)),
                            "total_equity": _safe_int(row.get("ACCT_657", 0)),
                        })

            # Parse FS220L.txt - commercial loan detail (CRE breakdown)
            fs220l_files = [f for f in file_list if "FS220L" in f.upper()]
            if fs220l_files:
                print(f"    Reading {fs220l_files[0]}...")
                with zf.open(fs220l_files[0]) as f:
                    content = f.read()
                    for encoding in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            text = content.decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    else:
                        text = content.decode("latin-1")

                    reader = csv.DictReader(io.StringIO(text))
                    if reader.fieldnames:
                        reader.fieldnames = [h.strip() for h in reader.fieldnames]

                    for row in reader:
                        cu_num = (row.get("CU_NUMBER") or row.get("cu_number") or "").strip()
                        if not cu_num:
                            continue
                        if cu_num not in cu_data:
                            cu_data[cu_num] = {}

                        cu_data[cu_num].update({
                            "multifamily": _safe_int(row.get("ACCT_400M", 0)) + _safe_int(row.get("ACCT_400M1", 0)),
                            "owner_occ_nonfarm": _safe_int(row.get("ACCT_400H2", 0)) + _safe_int(row.get("ACCT_400H3", 0)),
                            "nonocc_nonfarm": _safe_int(row.get("ACCT_400J2", 0)) + _safe_int(row.get("ACCT_400J3", 0)),
                            "construction_commercial": _safe_int(row.get("ACCT_143B3", 0)) + _safe_int(row.get("ACCT_143B4", 0)),
                            "commercial_industrial": _safe_int(row.get("ACCT_400L2", 0)),
                        })

            # Also try FS220H.txt as fallback for older data
            fs220h_files = [f for f in file_list if "FS220H" in f.upper()]
            if fs220h_files:
                print(f"    Reading {fs220h_files[0]} (supplemental)...")
                with zf.open(fs220h_files[0]) as f:
                    content = f.read()
                    for encoding in ["utf-8", "latin-1", "cp1252"]:
                        try:
                            text = content.decode(encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                    else:
                        text = content.decode("latin-1")

                    reader = csv.DictReader(io.StringIO(text))
                    if reader.fieldnames:
                        reader.fieldnames = [h.strip() for h in reader.fieldnames]

                    for row in reader:
                        cu_num = (row.get("CU_NUMBER") or row.get("cu_number") or "").strip()
                        if not cu_num or cu_num not in cu_data:
                            continue
                        # Only fill in if not already populated from FS220L
                        if not cu_data[cu_num].get("multifamily"):
                            cu_data[cu_num].setdefault("owner_occ_nonfarm", _safe_int(row.get("ACCT_400H", 0)))
                            cu_data[cu_num].setdefault("nonocc_nonfarm", _safe_int(row.get("ACCT_400J", 0)))

    except zipfile.BadZipFile:
        print(f"  WARNING: Downloaded file is not a valid ZIP")
        return {}
    finally:
        # Clean up temp ZIP
        try:
            os.remove(zip_path)
        except Exception:
            pass

    print(f"  Parsed financial data for {len(cu_data):,} credit unions")

    # Cache the parsed data
    with open(cache_file, "w") as f:
        json.dump(cu_data, f)
    print(f"  Cached to {cache_file}")

    return cu_data


def _safe_int(val) -> int:
    """Safely convert a value to int, returning 0 on failure."""
    try:
        return int(float(str(val).strip().replace(",", ""))) if val else 0
    except (ValueError, TypeError):
        return 0


def search_ncua_credit_union(name: str, ncua_data: dict = None) -> dict:
    """
    Search NCUA for a credit union by name.
    First tries the NCUA REST API, then falls back to local cached data.
    Returns dict with charter number and basic info, or None.
    """
    clean = name.strip()
    if not clean:
        return None

    # ── Attempt 1: NCUA REST API search ──
    payload = {
        "searchText": clean,
        "rdSearchType": "cuname",
        "rdSearchRadiusList": None,
        "is_mainOffice": True,
        "is_mdi": False,
        "is_member": False,
        "is_drive": False,
        "is_atm": False,
        "is_shared": False,
        "is_bilingual": False,
        "is_credit_builder": False,
        "is_fin_counseling": False,
        "is_homebuyer": False,
        "is_school": False,
        "is_low_wire": False,
        "is_no_draft": False,
        "is_no_tax": False,
        "is_payday": False,
        "skip": 0,
        "take": 10,
        "sort_item": "",
        "sort_direction": "",
    }

    api_result = None
    try:
        resp = requests.post(
            NCUA_SEARCH_URL,
            json=payload,
            timeout=15,
            headers={"Content-Type": "application/json", "User-Agent": "CRE-Lender-Match/1.0"},
        )
        resp.raise_for_status()
        data = resp.json()
        results = data.get("list", [])

        if not results and len(clean.split()) > 3:
            # Retry with shorter name
            payload["searchText"] = " ".join(clean.split()[:3])
            resp = requests.post(NCUA_SEARCH_URL, json=payload, timeout=15,
                                 headers={"Content-Type": "application/json", "User-Agent": "CRE-Lender-Match/1.0"})
            resp.raise_for_status()
            results = resp.json().get("list", [])

        if results:
            # Score for best match
            name_upper = name.upper()
            best, best_score = None, 0
            for r in results:
                rname = (r.get("creditUnionName") or "").upper()
                name_words = set(name_upper.split())
                r_words = set(rname.split())
                score = len(name_words & r_words)
                if name_upper in rname or rname in name_upper:
                    score += 10
                if score > best_score:
                    best_score = score
                    best = r
            if best and best_score >= 1:
                api_result = {
                    "charter": best.get("creditUnionNumber"),
                    "name": best.get("creditUnionName", ""),
                    "city": best.get("city", ""),
                    "state": best.get("state", ""),
                    "zip": best.get("zipcode", ""),
                    "address": best.get("street", ""),
                    "phone": best.get("phone", ""),
                    "website": best.get("url", ""),
                }
    except Exception:
        pass

    if api_result:
        return api_result

    # ── Attempt 2: Search cached FOICU data (name fuzzy match) ──
    if ncua_data:
        name_upper = name.upper()
        # Normalize: remove common suffixes and punctuation
        normalized = (name_upper
                      .replace("FEDERAL CREDIT UNION", "")
                      .replace("CREDIT UNION", "")
                      .replace("FCU", "")
                      .replace(",", " ").replace(".", " ").replace("-", " ")
                      .strip())
        skip_words = {"THE", "OF", "AND", "INC", "LLC", "NA"}
        search_words = [w for w in normalized.split() if w not in skip_words and len(w) > 1]

        if not search_words:
            search_words = normalized.split()[:1]

        best_charter, best_score = None, 0
        for cu_num, cu_info in ncua_data.items():
            cu_name = (cu_info.get("cu_name") or "").upper()
            if not cu_name:
                continue

            # Exact substring match (highest priority)
            if normalized and normalized in cu_name:
                score = 100
            elif cu_name in name_upper:
                score = 100
            else:
                # Word-level matching
                cu_words = set(cu_name.replace(",", " ").replace(".", " ").split())
                score = sum(1 for w in search_words if w in cu_words)
                # Partial word match (e.g. "BETHPAGE" matches "BETHPAGE")
                if score == 0:
                    score = sum(1 for w in search_words if any(w in cw or cw in w for cw in cu_words if len(cw) > 2))

            if score > best_score and score >= 1:
                best_score = score
                best_charter = cu_num

        if best_charter:
            cu_info = ncua_data[best_charter]
            return {
                "charter": int(best_charter),
                "name": cu_info.get("cu_name", ""),
                "city": cu_info.get("city", ""),
                "state": cu_info.get("state", ""),
                "zip": "",
                "address": "",
                "phone": "",
                "website": "",
            }

    return None


def fetch_ncua_details(charter: int) -> dict:
    """Fetch detailed credit union info from NCUA by charter number."""
    try:
        resp = requests.get(
            f"{NCUA_DETAILS_URL}/{charter}",
            timeout=15,
            headers={"User-Agent": "CRE-Lender-Match/1.0"},
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("isError"):
            return None
        return data
    except Exception:
        return None


def compute_cu_cre_concentration(cu_number: str, ncua_data: dict) -> dict:
    """
    Compute CRE concentration for a credit union from NCUA call report data.
    Returns dict similar to compute_cre_concentration() for banks.
    """
    fin = ncua_data.get(str(cu_number))
    if not fin:
        return {}

    total_assets = fin.get("total_assets", 0) or 0
    total_equity = fin.get("total_equity", 0) or 0
    multifamily = fin.get("multifamily", 0) or 0
    owner_occ = fin.get("owner_occ_nonfarm", 0) or 0
    nonocc = fin.get("nonocc_nonfarm", 0) or 0
    construction = fin.get("construction_commercial", 0) or fin.get("construction", 0) or 0
    total_re = (fin.get("total_1st_mtg_re", 0) or 0) + (fin.get("total_other_re", 0) or 0)

    # CRE = construction + multifamily + non-owner-occupied commercial
    cre_total = construction + multifamily + nonocc

    cre_to_assets_pct = (cre_total / total_assets * 100) if total_assets > 0 else 0
    cre_to_capital_pct = (cre_total / total_equity * 100) if total_equity > 0 else 0
    mf_pct = (multifamily / total_re * 100) if total_re > 0 else 0

    return {
        "cre_total_k": cre_total,
        "cre_to_capital_pct": round(cre_to_capital_pct, 1),
        "cre_to_assets_pct": round(cre_to_assets_pct, 1),
        "multifamily_k": multifamily,
        "multifamily_pct_of_re": round(mf_pct, 1),
        "construction_k": construction,
        "nonfarm_nonres_k": owner_occ + nonocc,
        "nonfarm_nonocc_k": nonocc,
        "nonfarm_ownocc_k": owner_occ,
        "total_assets_k": total_assets,
        "total_equity_k": total_equity,
        "total_re_loans_k": total_re,
        "report_date": "",
    }


# ════════════════════════════════════════════════════════════════════════════════
# Data Processing: Aggregate and Rank Lenders
# ════════════════════════════════════════════════════════════════════════════════

def classify_institution_type(name: str, fdic_found: bool) -> str:
    """Classify institution as Bank, Credit Union, or Other."""
    name_upper = (name or "").upper()
    cu_keywords = ["CREDIT UNION", " FCU", " CU,", "FEDERAL CREDIT"]
    for kw in cu_keywords:
        if kw in name_upper:
            return "Credit Union"
    if fdic_found:
        return "Bank"
    # Check for common mortgage company patterns
    mc_keywords = ["MORTGAGE", "LENDING", "HOME LOANS", "CAPITAL", "FUNDING"]
    for kw in mc_keywords:
        if kw in name_upper:
            return "Mortgage Company"
    return "Other"


def aggregate_by_lender(loans: list, filers: dict, min_loan: int = None, max_loan: int = None) -> list:
    """
    Aggregate loan records by lender. Filter by loan amount.
    Returns sorted list of lender dicts.
    """
    lender_data = defaultdict(lambda: {
        "lei": "",
        "name": "",
        "deal_count": 0,
        "total_volume": 0,
        "loan_amounts": [],
        "years_active": set(),
        "counties": set(),
        "unit_ranges": defaultdict(int),
    })

    filtered_count = 0
    for loan in loans:
        try:
            loan_amount = int(float(loan.get("loan_amount", 0)))
        except (ValueError, TypeError):
            continue

        # Apply loan amount filters
        if min_loan and loan_amount < min_loan:
            filtered_count += 1
            continue
        if max_loan and loan_amount > max_loan:
            filtered_count += 1
            continue

        lei = loan.get("lei", "unknown")
        entry = lender_data[lei]
        entry["lei"] = lei
        entry["name"] = filers.get(lei, lei)
        entry["deal_count"] += 1
        entry["total_volume"] += loan_amount
        entry["loan_amounts"].append(loan_amount)

        year = loan.get("activity_year", "")
        if year:
            entry["years_active"].add(str(year))

        county = loan.get("county_code", "")
        if county:
            entry["counties"].add(county)

        units = loan.get("total_units", "")
        if units:
            entry["unit_ranges"][units] += 1

    if filtered_count > 0:
        print(f"  Filtered out {filtered_count:,} loans outside loan amount range")

    # Convert to sorted list
    rankings = []
    for lei, data in lender_data.items():
        amounts = data["loan_amounts"]
        avg_loan = sum(amounts) / len(amounts) if amounts else 0
        min_originated = min(amounts) if amounts else 0
        max_originated = max(amounts) if amounts else 0

        rankings.append({
            "lei": lei,
            "name": data["name"],
            "deal_count": data["deal_count"],
            "total_volume": data["total_volume"],
            "avg_loan_size": int(avg_loan),
            "min_loan_originated": min_originated,
            "max_loan_originated": max_originated,
            "years_active": sorted(data["years_active"]),
            "counties_served": len(data["counties"]),
            "unit_breakdown": dict(data["unit_ranges"]),
        })

    # Sort by deal count (primary), then total volume (secondary)
    rankings.sort(key=lambda x: (x["deal_count"], x["total_volume"]), reverse=True)

    # Add rank
    for i, r in enumerate(rankings, 1):
        r["rank"] = i

    print(f"  Found {len(rankings):,} unique lenders")
    return rankings


def enrich_lenders(rankings: list, max_enrich: int = 100, ncua_data: dict = None) -> list:
    """
    Enrich top lenders with FDIC (banks) and NCUA (credit unions) financial data.
    Adds CRE concentration, assets, website, location, CEO (for CUs).
    """
    count = min(len(rankings), max_enrich)

    # First pass: classify all lenders so we know which are CUs
    for lender in rankings:
        lender["institution_type"] = classify_institution_type(lender["name"], False)

    cu_count = sum(1 for r in rankings[:count] if r["institution_type"] == "Credit Union")
    bank_count = count - cu_count
    print(f"\n  Enriching top {count} lenders ({bank_count} banks via FDIC, {cu_count} credit unions via NCUA)...")

    progress = tqdm(range(count), desc="  Enriching") if tqdm else range(count)

    for i in progress:
        lender = rankings[i]
        name = lender["name"]
        is_cu = lender["institution_type"] == "Credit Union"

        if is_cu:
            # ── Credit Union: use NCUA ──
            ncua_inst = search_ncua_credit_union(name, ncua_data=ncua_data)
            if ncua_inst:
                lender["ncua_match"] = True
                charter = ncua_inst.get("charter")
                lender["ncua_charter"] = charter
                lender["city"] = ncua_inst.get("city", "")
                lender["state"] = ncua_inst.get("state", "")
                lender["address"] = ncua_inst.get("address", "")
                lender["zip"] = ncua_inst.get("zip", "")
                lender["website"] = ncua_inst.get("website", "")
                lender["phone"] = ncua_inst.get("phone", "")

                # Get detailed info (assets, CEO, member count)
                if charter:
                    details = fetch_ncua_details(charter)
                    if details:
                        assets_str = details.get("creditUnionAssets", "0")
                        try:
                            assets_val = int(str(assets_str).replace(",", ""))
                        except (ValueError, TypeError):
                            assets_val = 0
                        lender["total_assets_m"] = round(assets_val / 1_000_000, 1)
                        lender["ceo"] = details.get("creditUnionCeo", "")
                        lender["member_count"] = details.get("creditUnionNom", "")
                        lender["peer_group"] = details.get("creditUnionPeerGroup", "")
                        # Prefer website from details if available
                        if details.get("creditUnionWebsite"):
                            lender["website"] = details["creditUnionWebsite"]

                    # Get CRE concentration from cached call report data
                    if ncua_data:
                        cre = compute_cu_cre_concentration(str(charter), ncua_data)
                        if cre:
                            lender.update(cre)
            else:
                lender["ncua_match"] = False

            lender["fdic_match"] = False
            time.sleep(NCUA_SEARCH_DELAY)

        else:
            # ── Bank / Other: use FDIC ──
            inst = search_fdic_institution(name)
            if inst:
                lender["fdic_match"] = True
                lender["fdic_cert"] = inst.get("CERT", "")
                lender["fdic_name"] = inst.get("NAME", "")
                lender["city"] = inst.get("CITY", "")
                lender["state"] = inst.get("STALP", "")
                lender["address"] = inst.get("ADDRESS", "")
                lender["zip"] = inst.get("ZIP", "")
                lender["website"] = inst.get("WEBADDR", "")
                lender["total_assets_m"] = round((inst.get("ASSET", 0) or 0) / 1000, 1)
                lender["offices"] = inst.get("OFFDOM", "")
                lender["specialization"] = inst.get("SPECGRPN", "")

                cert = inst.get("CERT")
                if cert:
                    fin = fetch_fdic_financials(cert)
                    cre = compute_cre_concentration(fin)
                    lender.update(cre)

                # Reclassify based on FDIC result
                lender["institution_type"] = classify_institution_type(name, True)
            else:
                lender["fdic_match"] = False

            lender["ncua_match"] = False
            time.sleep(FDIC_SEARCH_DELAY)

        if not tqdm:
            if (i + 1) % 10 == 0:
                print(f"    ...enriched {i + 1}/{count}")

    # Classify remaining (un-enriched) lenders
    for lender in rankings[count:]:
        lender["fdic_match"] = False
        lender["ncua_match"] = False

    # Count results
    banks = sum(1 for r in rankings if r["institution_type"] == "Bank")
    cus = sum(1 for r in rankings if r["institution_type"] == "Credit Union")
    other = len(rankings) - banks - cus
    enriched_cus = sum(1 for r in rankings[:count] if r.get("ncua_match"))
    enriched_banks = sum(1 for r in rankings[:count] if r.get("fdic_match"))
    print(f"  Results: {banks} banks ({enriched_banks} enriched), {cus} credit unions ({enriched_cus} enriched), {other} other")

    return rankings


# ════════════════════════════════════════════════════════════════════════════════
# Excel Output
# ════════════════════════════════════════════════════════════════════════════════

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MONEY_FORMAT = '#,##0'
MONEY_FORMAT_M = '#,##0.0"M"'
PCT_FORMAT = '0.0"%"'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, num_cols):
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, min_width=10, max_width=40):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def fmt_dollars(amount):
    """Format dollar amount for display."""
    if amount >= 1_000_000:
        return f"${amount / 1_000_000:,.1f}M"
    elif amount >= 1_000:
        return f"${amount / 1_000:,.0f}K"
    return f"${amount:,.0f}"


def create_excel_output(rankings: list, loans: list, search_params: dict, filename: str):
    """Create formatted Excel workbook with multiple sheets."""
    wb = Workbook()

    # ── Sheet 1: Search Summary ──
    ws_summary = wb.active
    ws_summary.title = "Search Summary"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    summary_data = [
        ("CRE Lender Match Report", ""),
        ("", ""),
        ("Search Parameters", ""),
        ("Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Geography", search_params.get("geo_label", "")),
        ("Property Type", search_params.get("property_type", "Multifamily")),
        ("Loan Range", f"{fmt_dollars(search_params.get('min_loan', 0))} - {fmt_dollars(search_params.get('max_loan', 0))}" if search_params.get("min_loan") or search_params.get("max_loan") else "All"),
        ("Years", ", ".join(str(y) for y in search_params.get("years", []))),
        ("", ""),
        ("Results Summary", ""),
        ("Total Originations Found", f"{len(loans):,}"),
        ("Unique Lenders", f"{len(rankings):,}"),
        ("Banks", str(sum(1 for r in rankings if r.get("institution_type") == "Bank"))),
        ("Credit Unions", str(sum(1 for r in rankings if r.get("institution_type") == "Credit Union"))),
        ("Other Lenders", str(sum(1 for r in rankings if r.get("institution_type") not in ("Bank", "Credit Union")))),
        ("Total Volume", fmt_dollars(sum(r["total_volume"] for r in rankings))),
    ]
    for row_data in summary_data:
        ws_summary.append(row_data)

    # Style summary
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws_summary["A3"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws_summary["A10"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # ── Sheet 2: Lender Rankings ──
    ws_rank = wb.create_sheet("Lender Rankings")
    ws_rank.sheet_properties.tabColor = "2E75B6"

    rank_headers = [
        "Rank", "Lender Name", "Type", "Deals", "Total Volume",
        "Avg Loan Size", "Min Loan", "Max Loan", "Years Active",
        "CRE % of Assets", "Multifamily ($K)", "Construction ($K)",
        "Non-Owner-Occ ($K)", "Owner-Occ ($K)",
        "Total Assets ($M)", "City", "State",
        "Website", "CEO/Contact", "Data Source",
    ]
    ws_rank.append(rank_headers)
    style_header_row(ws_rank, len(rank_headers))

    for r in rankings:
        # Determine data source
        if r.get("fdic_match"):
            data_source = "FDIC"
        elif r.get("ncua_match"):
            data_source = "NCUA"
        else:
            data_source = ""

        ws_rank.append([
            r.get("rank", ""),
            r.get("name", ""),
            r.get("institution_type", ""),
            r.get("deal_count", 0),
            r.get("total_volume", 0),
            r.get("avg_loan_size", 0),
            r.get("min_loan_originated", 0),
            r.get("max_loan_originated", 0),
            ", ".join(r.get("years_active", [])),
            r.get("cre_to_assets_pct", ""),
            r.get("multifamily_k", ""),
            r.get("construction_k", ""),
            r.get("nonfarm_nonocc_k", ""),
            r.get("nonfarm_ownocc_k", ""),
            r.get("total_assets_m", ""),
            r.get("city", ""),
            r.get("state", ""),
            r.get("website", ""),
            r.get("ceo", ""),
            data_source,
        ])

    # Format money columns (Total Volume, Avg, Min, Max = cols 5-8)
    for row in ws_rank.iter_rows(min_row=2, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    # Format CRE breakdown columns (Multifamily, Construction, Non-Owner-Occ, Owner-Occ = cols 11-14)
    for row in ws_rank.iter_rows(min_row=2, min_col=11, max_col=14):
        for cell in row:
            cell.number_format = MONEY_FORMAT

    auto_width(ws_rank)

    # ── Sheet 3: Loan Details ──
    ws_loans = wb.create_sheet("Loan Details")
    ws_loans.sheet_properties.tabColor = "548235"

    loan_headers = [
        "Year", "Lender (LEI)", "Lender Name", "Loan Amount",
        "Total Units", "Property Value", "Interest Rate",
        "Loan Type", "Loan Purpose", "Lien Status",
        "County", "State", "Census Tract",
    ]
    ws_loans.append(loan_headers)
    style_header_row(ws_loans, len(loan_headers))

    # Map loan type codes
    loan_type_map = {"1": "Conventional", "2": "FHA", "3": "VA", "4": "USDA"}
    loan_purpose_map = {"1": "Purchase", "2": "Improvement", "31": "Refinance", "32": "Cash-Out Refi", "4": "Other"}
    lien_map = {"1": "First Lien", "2": "Subordinate"}

    filers_map = {loan.get("lei", ""): "" for loan in loans}  # Will be filled
    for loan in loans:
        lei = loan.get("lei", "")
        try:
            amount = int(float(loan.get("loan_amount", 0)))
        except (ValueError, TypeError):
            amount = 0

        try:
            prop_val = int(float(loan.get("property_value", 0)))
        except (ValueError, TypeError):
            prop_val = ""

        ws_loans.append([
            loan.get("activity_year", ""),
            lei,
            next((r["name"] for r in rankings if r["lei"] == lei), lei),
            amount,
            loan.get("total_units", ""),
            prop_val,
            loan.get("interest_rate", ""),
            loan_type_map.get(loan.get("loan_type", ""), loan.get("loan_type", "")),
            loan_purpose_map.get(loan.get("loan_purpose", ""), loan.get("loan_purpose", "")),
            lien_map.get(str(loan.get("lien_status", "")), loan.get("lien_status", "")),
            loan.get("county_code", ""),
            loan.get("state_code", ""),
            loan.get("census_tract", ""),
        ])

    for row in ws_loans.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = MONEY_FORMAT

    auto_width(ws_loans)

    # ── Sheet 4: Salesforce Import ──
    ws_sf = wb.create_sheet("Salesforce Import")
    ws_sf.sheet_properties.tabColor = "00A1E0"

    sf_headers = [
        "Account Name", "Type", "Industry", "Website", "Phone",
        "BillingStreet", "BillingCity", "BillingState", "BillingPostalCode",
        "Description", "CRE_Concentration__c", "Total_Assets__c",
        "MF_Deal_Count__c", "MF_Volume__c", "Avg_Loan_Size__c",
    ]
    ws_sf.append(sf_headers)
    style_header_row(ws_sf, len(sf_headers))

    for r in rankings:
        if not r.get("name"):
            continue
        inst_type = r.get("institution_type", "Other")
        sf_type = "Bank" if inst_type == "Bank" else "Credit Union" if inst_type == "Credit Union" else "Other"

        description = (
            f"Active CRE lender - {r['deal_count']} multifamily deals "
            f"({', '.join(r.get('years_active', []))}). "
            f"Avg loan size: {fmt_dollars(r.get('avg_loan_size', 0))}. "
            f"Range: {fmt_dollars(r.get('min_loan_originated', 0))} - {fmt_dollars(r.get('max_loan_originated', 0))}."
        )

        ws_sf.append([
            r.get("name", ""),
            sf_type,
            "Banking" if inst_type in ("Bank", "Credit Union") else "Financial Services",
            r.get("website", ""),
            r.get("phone", ""),
            r.get("address", ""),
            r.get("city", ""),
            r.get("state", ""),
            r.get("zip", ""),
            description,
            r.get("cre_to_assets_pct", ""),
            r.get("total_assets_m", ""),
            r.get("deal_count", 0),
            r.get("total_volume", 0),
            r.get("avg_loan_size", 0),
        ])

    auto_width(ws_sf)

    # ── Sheet 5: ActiveCampaign Import ──
    ws_ac = wb.create_sheet("ActiveCampaign Import")
    ws_ac.sheet_properties.tabColor = "356AE6"

    ac_headers = [
        "Email", "First Name", "Last Name", "Organization",
        "Phone", "Tags",
    ]
    ws_ac.append(ac_headers)
    style_header_row(ws_ac, len(ac_headers))

    for r in rankings:
        if not r.get("name"):
            continue

        # Build smart tags
        tags = []
        inst_type = r.get("institution_type", "Other")
        tags.append(inst_type)

        deal_count = r.get("deal_count", 0)
        if deal_count >= 20:
            tags.append("High Activity Lender")
        elif deal_count >= 5:
            tags.append("Active Lender")
        else:
            tags.append("Occasional Lender")

        cre_pct = r.get("cre_to_assets_pct", 0) or 0
        if cre_pct >= 30:
            tags.append("High CRE Concentration")
        elif cre_pct >= 15:
            tags.append("Moderate CRE Concentration")

        assets_m = r.get("total_assets_m", 0) or 0
        if assets_m >= 10000:
            tags.append("Large Bank ($10B+)")
        elif assets_m >= 1000:
            tags.append("Mid-Size Bank ($1-10B)")
        elif assets_m >= 100:
            tags.append("Community Bank ($100M-$1B)")
        elif assets_m > 0:
            tags.append("Small Bank (<$100M)")

        tags.append("Multifamily Lender")
        tags.append(f"CRE Scraper - {search_params.get('geo_label', '')}")

        ws_ac.append([
            "",  # Email - to be filled
            "",  # First Name
            "",  # Last Name
            r.get("name", ""),
            "",  # Phone
            ", ".join(tags),
        ])

    auto_width(ws_ac)

    # Save
    wb.save(filename)
    print(f"\n  Saved: {filename}")
    return filename


# ════════════════════════════════════════════════════════════════════════════════
# Main CLI
# ════════════════════════════════════════════════════════════════════════════════

def print_banner():
    print("=" * 70)
    print("  CRE Lender Match Tool")
    print("  Find banks & credit unions actively lending in your market")
    print("=" * 70)


def print_top_results(rankings: list, top_n: int = 20, portfolio_mode: bool = False, property_type: str = ""):
    """Print top lenders to console."""
    print(f"\n{'=' * 70}")
    if portfolio_mode:
        label = CRE_TYPE_MAP.get(property_type, {}).get("label", "CRE")
        print(f"  TOP {min(top_n, len(rankings))} LENDERS BY {label.upper()} PORTFOLIO")
        print(f"{'=' * 70}")
        print(f"  {'#':<4} {'Lender':<35} {'Type':<10} {'Portfolio $M':<14} {'CRE %':<8} {'Assets $M':<12}")
        print(f"  {'-'*4} {'-'*35} {'-'*10} {'-'*14} {'-'*8} {'-'*12}")

        for r in rankings[:top_n]:
            name = r["name"][:33]
            inst_type = r.get("institution_type", "")[:9]
            portfolio_m = r.get("type_portfolio_m", 0)
            cre_pct = r.get("cre_to_assets_pct", 0) or 0
            assets_m = r.get("total_assets_m", 0) or 0
            print(
                f"  {r['rank']:<4} {name:<35} {inst_type:<10} "
                f"{'$' + str(portfolio_m) + 'M':<14} {str(cre_pct) + '%':<8} {'$' + str(assets_m) + 'M':<12}"
            )
    else:
        print(f"  TOP {min(top_n, len(rankings))} LENDERS BY DEAL COUNT")
        print(f"{'=' * 70}")
        print(f"  {'#':<4} {'Lender':<40} {'Type':<10} {'Deals':<6} {'Volume':<14} {'Avg Size':<12}")
        print(f"  {'-'*4} {'-'*40} {'-'*10} {'-'*6} {'-'*14} {'-'*12}")

        for r in rankings[:top_n]:
            name = r["name"][:38]
            inst_type = r.get("institution_type", "")[:9]
            print(
                f"  {r['rank']:<4} {name:<40} {inst_type:<10} "
                f"{r['deal_count']:<6} {fmt_dollars(r['total_volume']):<14} "
                f"{fmt_dollars(r['avg_loan_size']):<12}"
            )

    print(f"{'=' * 70}")


def main():
    parser = argparse.ArgumentParser(
        description="Find banks & credit unions actively lending on CRE in your market",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --city Miami --state FL --min-loan 2000000 --max-loan 10000000
  %(prog)s --city Phoenix --state AZ --years 2022,2023,2024
  %(prog)s --county 12086 --min-loan 1000000
  %(prog)s --state TX --min-loan 5000000 --max-loan 25000000
  %(prog)s --msa 33100 --years 2023,2024 --max-enrich 50
        """,
    )

    # Geography
    geo_group = parser.add_argument_group("Geography (pick one approach)")
    geo_group.add_argument("--city", help="City name (use with --state)")
    geo_group.add_argument("--state", help="State abbreviation (e.g., FL, TX, CA)")
    geo_group.add_argument("--county", help="County FIPS code (e.g., 12086 for Miami-Dade)")
    geo_group.add_argument("--msa", help="MSA/MD code (e.g., 33100 for Miami)")

    # Filters
    filter_group = parser.add_argument_group("Filters")
    filter_group.add_argument("--min-loan", type=int, default=None, help="Minimum loan amount in dollars")
    filter_group.add_argument("--max-loan", type=int, default=None, help="Maximum loan amount in dollars")
    filter_group.add_argument(
        "--years", default=",".join(str(y) for y in DEFAULT_YEARS),
        help=f"Comma-separated years (default: {','.join(str(y) for y in DEFAULT_YEARS)})",
    )
    filter_group.add_argument(
        "--property-type",
        choices=[
            "all-cre", "multifamily", "student-housing", "affordable-housing",
            "senior-housing", "manufactured-housing", "hospitality", "office",
            "retail", "industrial", "self-storage", "healthcare", "senior-care",
            "mixed-use", "special-purpose", "non-res", "construction", "land",
            "owner-occ", "sba-owner-user",
        ],
        default="multifamily",
        help="CRE property type. multifamily uses HMDA origination data. Others use FDIC/NCUA portfolio data.",
    )
    filter_group.add_argument(
        "--min-assets", type=int, default=50,
        help="Min bank assets in $M for portfolio search (default: 50)",
    )

    # Output
    out_group = parser.add_argument_group("Output")
    out_group.add_argument("-o", "--output", help="Output Excel filename (auto-generated if not specified)")
    out_group.add_argument("--max-enrich", type=int, default=100, help="Max lenders to enrich with FDIC data (default: 100)")
    out_group.add_argument("--no-enrich", action="store_true", help="Skip FDIC/NCUA enrichment (faster)")
    out_group.add_argument("--no-ncua-financials", action="store_true", help="Skip NCUA call report download (use basic CU info only)")
    out_group.add_argument("--top", type=int, default=25, help="Number of top results to display (default: 25)")

    args = parser.parse_args()

    print_banner()

    # ── Resolve Geography ──
    years = [int(y.strip()) for y in args.years.split(",")]
    geo_type = None
    geo_code = None
    geo_label = ""

    if args.county:
        geo_type = "counties"
        geo_code = args.county
        geo_label = f"County FIPS {args.county}"
    elif args.msa:
        geo_type = "msamds"
        geo_code = args.msa
        geo_label = f"MSA {args.msa}"
    elif args.city and args.state:
        print("\nStep 1: Resolving geography...")
        result = geocode_city_to_fips(args.city, args.state)
        if result:
            geo_type = "counties"
            geo_code = result["county_fips"]
            geo_label = f"{args.city}, {args.state} ({result['county_name']} County, FIPS {geo_code})"
        else:
            # Fall back to state-level search
            print(f"  Falling back to state-level search for {args.state}")
            geo_type = "states"
            geo_code = args.state
            geo_label = f"State: {args.state}"
    elif args.state:
        geo_type = "states"
        geo_code = args.state
        geo_label = f"State: {args.state}"
    else:
        parser.error("Provide --city + --state, --state, --county, or --msa")

    print(f"\n  Geography: {geo_label}")
    print(f"  Years: {', '.join(str(y) for y in years)}")
    if args.min_loan:
        print(f"  Min loan: {fmt_dollars(args.min_loan)}")
    if args.max_loan:
        print(f"  Max loan: {fmt_dollars(args.max_loan)}")

    # ── Branch by Property Type ──
    property_type = args.property_type.replace("-", "_")

    if property_type != "multifamily":
        # ═══════════════════════════════════════════════════════════════════
        # Portfolio Search Mode (Construction, Office/Retail, Owner-Occ, All CRE)
        # Uses FDIC + NCUA call report portfolio data instead of HMDA originations
        # ═══════════════════════════════════════════════════════════════════
        state = args.state
        if not state:
            parser.error("Portfolio search (non-multifamily property types) requires --state")

        # Load NCUA data
        ncua_data = None
        if not args.no_ncua_financials:
            print("\nStep 2: Loading NCUA credit union financial data...")
            try:
                year, month = get_ncua_current_cycle()
                ncua_data = download_ncua_call_reports(year, month)
            except Exception as e:
                print(f"  WARNING: Could not load NCUA data: {e}")

        # Fetch CRE portfolio lenders
        print(f"\nStep 3: Fetching CRE portfolio data...")
        rankings = fetch_state_cre_lenders(
            state=state, property_type=property_type,
            min_assets_m=args.min_assets, ncua_data=ncua_data,
        )
        loans = []  # No individual loans in portfolio mode

        if not rankings:
            print("\n  No CRE lenders found for this search.")
            print("  Try lowering --min-assets or broadening to --property-type all-cre")
            sys.exit(0)

        # Display results
        print_top_results(rankings, top_n=args.top, portfolio_mode=True, property_type=property_type)

        search_params = {
            "geo_label": geo_label,
            "property_type": CRE_TYPE_MAP.get(property_type, {}).get("label", property_type),
            "min_loan": 0,
            "max_loan": 0,
            "years": [],
        }

    else:
        # ═══════════════════════════════════════════════════════════════════
        # HMDA Origination Search Mode (Multifamily)
        # Uses actual loan-level origination data from HMDA
        # ═══════════════════════════════════════════════════════════════════
        print("\nStep 2: Fetching lender institution directory...")
        filers = fetch_hmda_filers(years)

        print("\nStep 3: Downloading origination data...")
        loans = fetch_hmda_originations(geo_type, geo_code, years)

        if not loans:
            print("\n  No origination data found for this search.")
            print("  Try broadening your search (larger geography, wider loan range, more years).")
            sys.exit(0)

        # Aggregate by lender
        print("\nStep 4: Aggregating by lender...")
        rankings = aggregate_by_lender(loans, filers, args.min_loan, args.max_loan)

        if not rankings:
            print("\n  No lenders matched after filtering.")
            print("  Try widening your loan amount range.")
            sys.exit(0)

        # Enrich with FDIC + NCUA
        if not args.no_enrich:
            ncua_data = None
            if not args.no_ncua_financials:
                print("\nStep 5a: Loading NCUA credit union financial data...")
                try:
                    year, month = get_ncua_current_cycle()
                    ncua_data = download_ncua_call_reports(year, month)
                except Exception as e:
                    print(f"  WARNING: Could not load NCUA data: {e}")
                    print("  Credit unions will still be enriched with basic info (assets, website, CEO)")

            print("\nStep 5b: Enriching lenders with FDIC + NCUA data...")
            rankings = enrich_lenders(rankings, max_enrich=args.max_enrich, ncua_data=ncua_data)
        else:
            print("\nStep 5: Skipping enrichment (--no-enrich)")
            for r in rankings:
                r["institution_type"] = classify_institution_type(r["name"], False)

        # Display results
        print_top_results(rankings, top_n=args.top)

        search_params = {
            "geo_label": geo_label,
            "property_type": "Multifamily (5+ units)",
            "min_loan": args.min_loan or 0,
            "max_loan": args.max_loan or 0,
            "years": years,
        }

    # ── Generate Output (both modes) ──
    if not args.output:
        geo_short = geo_code.replace(",", "-") if geo_code else (args.state or "search")
        date_str = datetime.now().strftime("%Y%m%d")
        type_tag = f"_{property_type}" if property_type != "multifamily" else ""
        loan_range = ""
        if args.min_loan or args.max_loan:
            lo = f"{(args.min_loan or 0) // 1_000_000}M" if (args.min_loan or 0) >= 1_000_000 else f"{(args.min_loan or 0) // 1000}K"
            hi = f"{(args.max_loan or 0) // 1_000_000}M" if (args.max_loan or 0) >= 1_000_000 else f"{(args.max_loan or 0) // 1000}K"
            loan_range = f"_{lo}-{hi}"
        args.output = f"lender_match_{geo_short}{type_tag}{loan_range}_{date_str}.xlsx"

    print(f"\nGenerating Excel report...")
    create_excel_output(rankings, loans, search_params, args.output)

    # ── Summary ──
    banks = sum(1 for r in rankings if r.get("institution_type") == "Bank")
    cus = sum(1 for r in rankings if r.get("institution_type") == "Credit Union")
    total_vol = sum(r.get("total_volume", 0) for r in rankings)

    print(f"\n{'=' * 70}")
    print(f"  DONE!")
    print(f"  {len(rankings):,} lenders found ({banks} banks, {cus} credit unions)")
    if loans:
        print(f"  {len(loans):,} total originations, {fmt_dollars(total_vol)} total volume")
    else:
        print(f"  Total CRE portfolio: {fmt_dollars(total_vol)}")
    print(f"  Report saved to: {args.output}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
